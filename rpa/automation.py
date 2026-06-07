"""
RPA - Agendador de tarefas automáticas
Verifica empréstimos atrasados e envia notificações via schedule.
"""

import schedule
import time
import logging
from datetime import date

logging.basicConfig(
    filename='rpa/rpa_log.txt',
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s'
)


def verificar_atrasos():
    """Verifica empréstimos atrasados e notifica usuários."""
    try:
        import sys
        import os
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from app import app, Emprestimo, db
        from rpa.send_email import enviar_email_atraso
        from rpa.send_whatsapp import enviar_whatsapp_atraso

        with app.app_context():
            atrasados = Emprestimo.query.filter(
                Emprestimo.status == 'ativo',
                Emprestimo.data_devolucao_prevista < date.today()
            ).all()

            logging.info(f'Verificação de atrasos: {len(atrasados)} empréstimo(s) atrasado(s)')

            for emp in atrasados:
                emp.status = 'atrasado'
                dados = emp.to_dict()

                # Notifica por e-mail
                resultado_email = enviar_email_atraso(dados)
                logging.info(f'E-mail atraso | Emprestimo #{emp.id} | {resultado_email}')

                # Notifica por WhatsApp se tiver telefone
                if emp.usuario and emp.usuario.telefone:
                    resultado_wpp = enviar_whatsapp_atraso(dados, emp.usuario.telefone)
                    logging.info(f'WhatsApp atraso | Emprestimo #{emp.id} | {resultado_wpp}')

            db.session.commit()
            print(f'[RPA] {len(atrasados)} notificação(ões) de atraso enviada(s).')

    except Exception as e:
        logging.error(f'FALHA | verificar_atrasos | {e}')
        print(f'[RPA] Erro: {e}')


def exportar_relatorio_excel():
    """Exporta relatório de empréstimos para Excel usando openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import sys, os
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from app import app, Emprestimo

        with app.app_context():
            emprestimos = Emprestimo.query.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Empréstimos'

        # Cabeçalho
        cabecalhos = ['ID', 'Livro', 'Autor', 'Usuário', 'E-mail', 'Empréstimo', 'Devolução Prevista', 'Status']
        for col, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=cab)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Dados
        cores_status = {'ativo': 'E8F5E9', 'devolvido': 'F5F5F5', 'atrasado': 'FFEBEE'}
        for row, emp in enumerate(emprestimos, 2):
            dados = [
                emp.id, emp.livro.titulo if emp.livro else '-',
                emp.livro.autor if emp.livro else '-',
                emp.usuario.nome if emp.usuario else '-',
                emp.usuario.email if emp.usuario else '-',
                str(emp.data_emprestimo), str(emp.data_devolucao_prevista), emp.status
            ]
            cor = cores_status.get(emp.status, 'FFFFFF')
            for col, valor in enumerate(dados, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        # Ajusta largura das colunas
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        nome_arquivo = f'rpa/relatorio_emprestimos_{date.today()}.xlsx'
        wb.save(nome_arquivo)
        logging.info(f'SUCESSO | Relatório exportado: {nome_arquivo}')
        print(f'[RPA] Relatório gerado: {nome_arquivo}')
        return nome_arquivo

    except Exception as e:
        logging.error(f'FALHA | exportar_relatorio | {e}')
        print(f'[RPA] Erro ao exportar: {e}')
        return None


def notificar_desktop(titulo: str, mensagem: str):
    """Exibe notificação no sistema operacional via plyer."""
    try:
        from plyer import notification
        notification.notify(
            title=titulo,
            message=mensagem,
            app_name='Biblioteca Digital ADS',
            timeout=10
        )
        logging.info(f'SUCESSO | Notificação desktop: {titulo}')
    except Exception as e:
        logging.warning(f'Notificação desktop indisponível: {e}')


if __name__ == '__main__':
    print('[RPA] Agendador iniciado. Verificações diárias às 08:00.')
    print('[RPA] Exportação de relatório toda sexta às 17:00.')
    print('[RPA] Pressione Ctrl+C para encerrar.\n')

    # Agenda tarefas
    schedule.every().day.at('08:00').do(verificar_atrasos)
    schedule.every().friday.at('17:00').do(exportar_relatorio_excel)

    # Executa uma vez ao iniciar para teste
    verificar_atrasos()
    exportar_relatorio_excel()
    notificar_desktop('Biblioteca Digital', 'Agendador RPA iniciado com sucesso!')

    while True:
        schedule.run_pending()
        time.sleep(60)
